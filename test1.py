### ping送信先リスト（Excelファイル）を読み込む ###
from datetime import date
import openpyxl
from ping3 import ping

# 調査対象（Excelファイル）を指定
wb_source=openpyxl.load_workbook('test1.xlsx')
ws_source=wb_source['Sheet1']

# レポート（Excelファイル）のファイル名を指定
filename_report='report_' + date.today().strftime('%Y%m%d') + '.xlsx'
openpyxl.Workbook().save(filename_report)

# データを2次元配列に格納
PingList=[[0 for i in range(ws_source.max_column)] for j in range(ws_source.max_row)]
for x in range(0, ws_source.max_row):
    for y in range(0, ws_source.max_column):
        PingList[x][y]=ws_source.cell(row=x+1, column=y+1).value

### ping一斉送信 ###
wb_report=openpyxl.load_workbook(filename_report)
ws_report=wb_report['Sheet']

#ヘッダーをセルに書き込み
ws_report.cell(row=1, column=1).value = 'IPアドレス'

for i in range(0, ws_source.max_row):        
    # ping送信先IPアドレスをセルに書き込み
    ws_report.cell(row=i+2, column=1).value=PingList[i][0]

    # ping応答時間をセルに書き込み（j回送信）
    for j in range(4):
        result=ping(PingList[i][0], unit='ms')
        ws_report.cell(row=1, column=j+2).value = str(j+1) + '回目'
        ws_report.cell(row=i+2, column=j+2).value=result
        print(result)

### 列幅を変更 ###
ws_report.column_dimensions['A'].width=15

### レポート（Excelファイル）に保存 ###
wb_report.save(filename_report)