import datetime
from os import mkdir
import openpyxl
from ping3 import ping
import requests

# ★1　調査対象ファイルを読み込む
with open('destinations.txt', 'r') as f:
    destinations=f.readlines()

# 調査結果ファイルを作成
try:
    #フォルダ作成
    mkdir('reports')
except:
    #既に存在する場合は何もしない
    pass
now=datetime.datetime.now().strftime('%Y%m%d%H%M%S')
save_file='reports/report_' + now + '.xlsx'
openpyxl.Workbook().save(save_file)

# シート名変更
wb=openpyxl.load_workbook(save_file)
ws=wb['Sheet']
ws.title='Report' + now

# ヘッダ情報をセルに書き込む
ws.cell(row=1, column=1).value='宛先'
ws.cell(row=1, column=2).value='応答時間'
ws.cell(row=1, column=3).value='HTTPレスポンス'

# ★1で読み込んだ行数でループ開始
for i, destination in enumerate(destinations):
    # 改行コード削除
    destination=destination.rstrip()
    # 宛先をセルに書き込む
    ws.cell(row=i+2, column=1).value=destination
    # ping送信
    result_ping=ping(destination, unit='ms')
    ws.cell(row=i+2, column=2).value=result_ping
    # HTTPリクエスト送信
    try:
        result_http=requests.get('http://' + destination, timeout=3.0).status_code
    except:
        result_http='エラー'
    ws.cell(row=i+2, column=3).value=result_http
    
# 調査結果ファイルを保存
wb.save(save_file)
wb.close()